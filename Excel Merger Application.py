#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Merger  /  Excel 合并工具
================================
A production-quality desktop application for merging Excel and CSV files.
Supports multiple merge modes, directions, progress tracking, and logging.

Requirements:
    pip install pandas openpyxl

Author: Senior Python Engineer
Version: 1.0.0
"""

import os
import re
import sys
import threading
import traceback
from pathlib import Path
from typing import Callable, Optional

# ── dependency check ──────────────────────────────────────────────────────────
try:
    import pandas as pd
    import openpyxl  # noqa – required by pandas ExcelWriter
except ImportError as _e:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(
        "Missing Dependency",
        f"Required package not found:\n{_e}\n\n"
        "Please run:\n  pip install pandas openpyxl"
    )
    sys.exit(1)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# ═════════════════════════════════════════════════════════════════════════════
#  DESIGN TOKENS
# ═════════════════════════════════════════════════════════════════════════════

APP_TITLE   = "Excel Merger  /  Excel 合并工具"
APP_VERSION = "1.0.0"

# Notion-inspired light palette
C = {
    "bg":           "#F7F7F5",   # page background
    "surface":      "#FFFFFF",   # card / input background
    "surface_alt":  "#F0EFED",   # hover / alternate row
    "border":       "#E3E2DF",   # subtle borders
    "border_focus": "#9B9B9B",   # focused border
    "accent":       "#2383E2",   # primary blue
    "accent_hover": "#1A6CC7",
    "accent_dim":   "#EBF3FD",   # tinted background
    "success":      "#0F9D58",
    "warning":      "#E8A000",
    "error":        "#D93025",
    "txt":          "#1A1A1A",   # primary text
    "txt_sec":      "#555555",   # secondary text
    "txt_mute":     "#999999",   # muted / placeholder
    "log_bg":       "#FBFAF8",   # log panel background
    "progress":     "#2383E2",
    "progress_bg":  "#E3E2DF",
    "white":        "#FFFFFF",
    "selection":    "#DBEAFE",
}

F = {
    "title":   ("Segoe UI Semibold", 14),
    "heading": ("Segoe UI Semibold", 10),
    "body":    ("Segoe UI",          10),
    "small":   ("Segoe UI",           9),
    "caption": ("Segoe UI",           8),
    "mono":    ("Consolas",           9),
    "btn":     ("Segoe UI Semibold", 10),
    "btn_big": ("Segoe UI Semibold", 11),
}

LARGE_FILE_WARN_MB = 100   # warn when total input exceeds this


# ═════════════════════════════════════════════════════════════════════════════
#  DATA PROCESSOR  (pure data logic, zero UI references)
# ═════════════════════════════════════════════════════════════════════════════

class DataProcessor:
    """All file I/O and merge operations. Completely decoupled from the UI."""

    # ─── reading ─────────────────────────────────────────────────────────────

    @staticmethod
    def read_file(filepath: str) -> dict[str, pd.DataFrame]:
        """
        Read an Excel or CSV file.
        Returns {sheet_name: DataFrame} dict.
        Raises ValueError / Exception on failure.
        """
        ext = Path(filepath).suffix.lower()
        name = Path(filepath).name

        if ext == ".csv":
            # try common encodings
            for enc in ("utf-8", "utf-8-sig", "gbk", "gb18030", "latin-1"):
                try:
                    df = pd.read_csv(filepath, encoding=enc, on_bad_lines="skip")
                    return {Path(filepath).stem: df}
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"Could not decode '{name}' with any known encoding.")

        elif ext in (".xlsx", ".xls"):
            engine = "openpyxl" if ext == ".xlsx" else "xlrd"
            try:
                dfs = pd.read_excel(filepath, sheet_name=None, engine=engine)
            except Exception:
                # fallback: try openpyxl for .xls too
                dfs = pd.read_excel(filepath, sheet_name=None, engine="openpyxl")
            if not dfs:
                raise ValueError(f"'{name}' contains no sheets.")
            return dfs

        else:
            raise ValueError(f"Unsupported file type: '{ext}'")

    # ─── sheet name helpers ───────────────────────────────────────────────────

    @staticmethod
    def safe_sheet_name(raw: str, used: list[str]) -> str:
        """
        Produce an Excel-legal sheet name (≤31 chars, no illegal chars)
        that is unique within *used*.
        """
        # Excel forbids: \ / * ? : [ ]
        clean = re.sub(r'[\\/*?:\[\]]', "_", str(raw)).strip()[:28] or "Sheet"
        if clean not in used:
            return clean
        for i in range(2, 9999):
            candidate = f"{clean[:25]}_{i}"
            if candidate not in used:
                return candidate
        return clean + "_X"

    # ─── merge operations ─────────────────────────────────────────────────────

    @staticmethod
    def merge_vertical(frames: list[pd.DataFrame]) -> pd.DataFrame:
        """Stack DataFrames row-wise, unioning columns."""
        if not frames:
            return pd.DataFrame()
        return pd.concat(frames, axis=0, ignore_index=True, sort=False)

    @staticmethod
    def merge_horizontal(
        frames: list[pd.DataFrame],
        key: str,
        how: str
    ) -> pd.DataFrame:
        """Join DataFrames side-by-side on *key* column."""
        if not frames:
            return pd.DataFrame()
        result = frames[0]
        for df in frames[1:]:
            # suffix duplicates with _2, _3, …
            dup_cols = [c for c in df.columns if c != key and c in result.columns]
            suffix = f"_{len([c for c in result.columns if c.startswith(key)]) + 1}"
            result = pd.merge(result, df, on=key, how=how,
                              suffixes=("", suffix if dup_cols else ""))
        return result

    # ─── batch loader ─────────────────────────────────────────────────────────

    @staticmethod
    def load_frames(
        files: list[str],
        progress_cb: Callable[[int, int, str], None],
        log_cb: Callable[[str, str], None]
    ) -> list[tuple[str, str, pd.DataFrame]]:
        """
        Load all sheets from all files.
        Returns list of (filepath, sheet_name, DataFrame).
        """
        all_frames: list[tuple[str, str, pd.DataFrame]] = []
        total = len(files)

        for i, fp in enumerate(files):
            fname = Path(fp).name
            log_cb(f"Reading  {fname}", "info")

            try:
                sheets = DataProcessor.read_file(fp)
                if not sheets:
                    log_cb(f"  ⚠  No data in {fname}", "warning")
                    continue
                for sname, df in sheets.items():
                    if df.empty:
                        log_cb(f"  ⚠  Sheet '{sname}' is empty – skipped", "warning")
                        continue
                    # drop fully-empty columns
                    df = df.dropna(axis=1, how="all")
                    all_frames.append((fp, sname, df))
                    log_cb(
                        f"  ✓  '{sname}'  →  "
                        f"{len(df):,} rows × {len(df.columns)} cols",
                        "success"
                    )
            except Exception as exc:
                log_cb(f"  ✗  {fname}: {exc}", "error")

            progress_cb(i + 1, total, fname)

        return all_frames


# ═════════════════════════════════════════════════════════════════════════════
#  CUSTOM WIDGETS
# ═════════════════════════════════════════════════════════════════════════════

class FlatButton(tk.Label):
    """
    A lightweight flat button built on tk.Label so we can control
    background colour precisely on every platform.
    """
    def __init__(self, parent, text, command=None,
                 bg=C["surface"], fg=C["txt"],
                 hover_bg=C["surface_alt"],
                 font=F["body"], padx=12, pady=6,
                 border_color=C["border"], **kwargs):
        super().__init__(parent, text=text, bg=bg, fg=fg,
                         font=font, cursor="hand2",
                         padx=padx, pady=pady,
                         relief="flat", **kwargs)
        self._bg = bg
        self._hover_bg = hover_bg
        self._command = command
        self._draw_border(border_color)
        self.bind("<Enter>",  lambda _: self.config(bg=self._hover_bg))
        self.bind("<Leave>",  lambda _: self.config(bg=self._bg))
        self.bind("<Button-1>", self._on_click)

    def _draw_border(self, color):
        self.config(highlightbackground=color, highlightthickness=1,
                    highlightcolor=color)

    def _on_click(self, _event=None):
        if self._command and str(self.cget("state")) != "disabled":
            self._command()

    def set_state(self, enabled: bool):
        if enabled:
            self.config(state="normal", cursor="hand2", fg=C["txt"])
        else:
            self.config(state="disabled", cursor="", fg=C["txt_mute"])


class AccentButton(FlatButton):
    """Primary call-to-action button."""
    def __init__(self, parent, text, command=None, **kwargs):
        super().__init__(
            parent, text=text, command=command,
            bg=C["accent"], fg=C["white"],
            hover_bg=C["accent_hover"],
            border_color=C["accent"],
            font=F["btn_big"], padx=20, pady=9,
            **kwargs
        )
        self._normal_bg   = C["accent"]
        self._hover_bg    = C["accent_hover"]
        self._disabled_bg = C["border"]
        self.bind("<Enter>",  lambda _: self.config(bg=self._hover_bg)
                              if str(self.cget("state")) != "disabled" else None)
        self.bind("<Leave>",  lambda _: self.config(bg=self._normal_bg)
                              if str(self.cget("state")) != "disabled" else None)

    def set_state(self, enabled: bool):
        if enabled:
            self.config(state="normal", cursor="hand2",
                        bg=self._normal_bg, fg=C["white"])
        else:
            self.config(state="disabled", cursor="",
                        bg=self._disabled_bg, fg=C["txt_mute"])


class SectionLabel(tk.Frame):
    """Section header with index badge + title text."""
    def __init__(self, parent, index: str, title: str, **kwargs):
        super().__init__(parent, bg=C["bg"], **kwargs)
        badge = tk.Label(self, text=index, bg=C["accent"], fg=C["white"],
                         font=F["caption"], width=3, pady=1)
        badge.pack(side="left", padx=(0, 6))
        tk.Label(self, text=title, bg=C["bg"], fg=C["txt"],
                 font=F["heading"]).pack(side="left")


class HRule(tk.Frame):
    """1-px horizontal rule."""
    def __init__(self, parent, color=C["border"], **kwargs):
        super().__init__(parent, bg=color, height=1, **kwargs)


class Card(tk.Frame):
    """White rounded-looking card frame."""
    def __init__(self, parent, **kwargs):
        super().__init__(parent,
                         bg=C["surface"],
                         highlightbackground=C["border"],
                         highlightthickness=1,
                         **kwargs)


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═════════════════════════════════════════════════════════════════════════════

class ExcelMergerApp(tk.Tk):
    """
    Root application window.
    Responsibilities: build UI, wire events, delegate data work to DataProcessor.
    """

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("980x780")
        self.minsize(820, 660)
        self.configure(bg=C["bg"])

        # ── state variables ────────────────────────────────────────────────
        self.file_list:        list[str]   = []
        self.merge_mode       = tk.StringVar(value="A")
        self.direction        = tk.StringVar(value="vertical")
        self.join_type        = tk.StringVar(value="outer")
        self.include_sub      = tk.BooleanVar(value=False)
        self.key_column       = tk.StringVar()
        self.group_column     = tk.StringVar()
        self.sheet_naming     = tk.StringVar(value="file_sheet")
        self.output_dir       = tk.StringVar()
        self.output_name      = tk.StringVar(value="merged_output")
        self._running         = False

        self._build_ttk_style()
        self._build_ui()
        self._on_mode_change()
        self._on_direction_change()

    # ─── ttk style (used for Treeview, Scrollbar, Progressbar) ───────────────

    def _build_ttk_style(self):
        s = ttk.Style(self)
        try:
            s.theme_use("clam")
        except Exception:
            pass

        s.configure(".", background=C["bg"], foreground=C["txt"],
                    font=F["body"], relief="flat")

        s.configure("TScrollbar",
                    troughcolor=C["bg"],
                    background=C["border"],
                    arrowcolor=C["txt_mute"],
                    relief="flat", borderwidth=0)
        s.map("TScrollbar", background=[("active", C["txt_mute"])])

        s.configure("Treeview",
                    background=C["surface"],
                    fieldbackground=C["surface"],
                    foreground=C["txt"],
                    font=F["small"],
                    rowheight=26,
                    relief="flat",
                    borderwidth=0)
        s.configure("Treeview.Heading",
                    background=C["surface_alt"],
                    foreground=C["txt_sec"],
                    font=F["small"],
                    relief="flat")
        s.map("Treeview",
              background=[("selected", C["selection"])],
              foreground=[("selected", C["txt"])])

        s.configure("Horizontal.TProgressbar",
                    troughcolor=C["progress_bg"],
                    background=C["progress"],
                    thickness=5,
                    relief="flat",
                    borderwidth=0)

        s.configure("TCombobox",
                    fieldbackground=C["surface"],
                    background=C["surface"],
                    foreground=C["txt"],
                    arrowcolor=C["txt_sec"],
                    relief="flat")
        s.map("TCombobox",
              fieldbackground=[("readonly", C["surface"])],
              selectbackground=[("readonly", C["selection"])],
              selectforeground=[("readonly", C["txt"])])

    # ─── master layout ────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── top bar ──
        topbar = tk.Frame(self, bg=C["white"],
                          highlightbackground=C["border"],
                          highlightthickness=1)
        topbar.pack(fill="x", side="top")

        tk.Label(topbar,
                 text="⊞",
                 font=("Segoe UI", 16), bg=C["white"],
                 fg=C["accent"]).pack(side="left", padx=(16, 6), pady=10)
        tk.Label(topbar,
                 text="Excel Merger",
                 font=F["title"], bg=C["white"],
                 fg=C["txt"]).pack(side="left", pady=10)
        tk.Label(topbar,
                 text=" /  Excel 合并工具",
                 font=("Segoe UI", 12), bg=C["white"],
                 fg=C["txt_sec"]).pack(side="left", pady=10)
        tk.Label(topbar,
                 text=f"v{APP_VERSION}",
                 font=F["caption"], bg=C["white"],
                 fg=C["txt_mute"]).pack(side="right", padx=16)

        # ── two-column body ──
        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", expand=True)

        # LEFT (scrollable settings)
        left_outer = tk.Frame(body, bg=C["bg"])
        left_outer.pack(side="left", fill="both", expand=True)

        # RIGHT (log panel, fixed width)
        right_outer = tk.Frame(body, bg=C["bg"], width=290)
        right_outer.pack(side="right", fill="both")
        right_outer.pack_propagate(False)

        self._build_settings_panel(left_outer)
        self._build_log_panel(right_outer)

    # ─── settings panel (scrollable) ─────────────────────────────────────────

    def _build_settings_panel(self, parent):
        canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0,
                           borderwidth=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        self._scroll_frame = tk.Frame(canvas, bg=C["bg"])

        self._scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)

        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # mousewheel scroll (Windows + macOS)
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        canvas.bind_all("<Button-4>",
            lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>",
            lambda e: canvas.yview_scroll( 1, "units"))

        sf = self._scroll_frame
        self._build_file_section(sf)
        self._build_merge_section(sf)
        self._build_output_section(sf)
        self._build_action_section(sf)

    # ─── §1 File Selection ────────────────────────────────────────────────────

    def _build_file_section(self, parent):
        SectionLabel(parent, "①", "File Selection  /  文件选择").pack(
            fill="x", padx=16, pady=(16, 6))

        card = Card(parent)
        card.pack(fill="x", padx=16, pady=(0, 10))

        # button row
        br = tk.Frame(card, bg=C["surface"])
        br.pack(fill="x", padx=12, pady=(10, 8))

        FlatButton(br, "➕  Add Files / 添加文件",
                   command=self._add_files).pack(side="left", padx=(0, 6))
        FlatButton(br, "📁  Add Folder / 添加文件夹",
                   command=self._add_folder).pack(side="left", padx=(0, 6))
        FlatButton(br, "✕  Clear All / 清空",
                   command=self._clear_files,
                   fg=C["error"]).pack(side="left")

        # subfolder checkbox (custom)
        sub_frame = tk.Frame(br, bg=C["surface"])
        sub_frame.pack(side="right")
        chk = tk.Checkbutton(sub_frame, text="Incl. Subfolders / 含子文件夹",
                              variable=self.include_sub,
                              bg=C["surface"], fg=C["txt_sec"],
                              activebackground=C["surface"],
                              font=F["small"],
                              selectcolor=C["surface"],
                              relief="flat", bd=0, cursor="hand2")
        chk.pack()

        # file list (Treeview)
        tree_frame = tk.Frame(card, bg=C["surface"])
        tree_frame.pack(fill="x", padx=12, pady=(0, 8))

        cols = ("name", "sheets", "size", "path")
        self.file_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            height=7, selectmode="extended"
        )
        for col, heading, w in [
            ("name",   "File Name / 文件名", 190),
            ("sheets", "Sheets / 表数",       70),
            ("size",   "Size / 大小",          80),
            ("path",   "Full Path / 路径",     280),
        ]:
            self.file_tree.heading(col, text=heading)
            self.file_tree.column(col, width=w, minwidth=50)

        vsb_tree = ttk.Scrollbar(tree_frame, orient="vertical",
                                 command=self.file_tree.yview)
        hsb_tree = ttk.Scrollbar(tree_frame, orient="horizontal",
                                 command=self.file_tree.xview)
        self.file_tree.configure(yscrollcommand=vsb_tree.set,
                                 xscrollcommand=hsb_tree.set)
        vsb_tree.pack(side="right", fill="y")
        hsb_tree.pack(side="bottom", fill="x")
        self.file_tree.pack(fill="x")

        # bottom row of file card
        bot = tk.Frame(card, bg=C["surface"])
        bot.pack(fill="x", padx=12, pady=(0, 10))
        FlatButton(bot, "✕  Remove Selected / 移除选中",
                   command=self._remove_selected,
                   fg=C["txt_sec"]).pack(side="left")
        self.file_count_lbl = tk.Label(
            bot, text="No files added  /  未添加文件",
            font=F["small"], bg=C["surface"], fg=C["txt_mute"]
        )
        self.file_count_lbl.pack(side="right")

    # ─── §2 Merge Settings ───────────────────────────────────────────────────

    def _build_merge_section(self, parent):
        SectionLabel(parent, "②", "Merge Settings  /  合并设置").pack(
            fill="x", padx=16, pady=(4, 6))

        card = Card(parent)
        card.pack(fill="x", padx=16, pady=(0, 10))

        inner = tk.Frame(card, bg=C["surface"])
        inner.pack(fill="x", padx=14, pady=12)

        # ── Merge Mode ──
        tk.Label(inner, text="Merge Mode  /  合并模式",
                 font=F["heading"], bg=C["surface"],
                 fg=C["txt"]).grid(row=0, column=0, columnspan=4,
                                   sticky="w", pady=(0, 8))

        mode_data = [
            ("A", "Mode A — All into One Sheet",
                  "模式 A — 合并成一张表"),
            ("B", "Mode B — Separate Sheets",
                  "模式 B — 分开成多个工作表"),
            ("C", "Mode C — Group by Column Value",
                  "模式 C — 按列值分组"),
        ]
        for row, (val, en, zh) in enumerate(mode_data, start=1):
            # colored indicator
            dot = tk.Label(inner,
                           text="◉" if self.merge_mode.get() == val else "○",
                           bg=C["surface"], fg=C["accent"],
                           font=F["body"], cursor="hand2")
            dot.grid(row=row, column=0, padx=(0, 6), pady=3, sticky="w")
            lbl = tk.Label(inner,
                           text=f"{en}  /  {zh}",
                           font=F["body"], bg=C["surface"],
                           fg=C["txt"], cursor="hand2")
            lbl.grid(row=row, column=1, columnspan=3, sticky="w", pady=3)
            # clicking label triggers radio
            for widget in (dot, lbl):
                widget.bind("<Button-1>",
                    lambda e, v=val: (
                        self.merge_mode.set(v),
                        self._on_mode_change(),
                        self._refresh_mode_dots(inner, mode_data)
                    ))

        self._mode_inner     = inner
        self._mode_data      = mode_data
        self._mode_dot_row   = 1

        # Mode B sub-options
        self.mode_b_row = tk.Frame(inner, bg=C["surface"])
        self.mode_b_row.grid(row=4, column=0, columnspan=4, sticky="w",
                             padx=20, pady=4)
        tk.Label(self.mode_b_row, text="Sheet Naming / 工作表命名:",
                 font=F["small"], bg=C["surface"],
                 fg=C["txt_sec"]).pack(side="left", padx=(0, 10))
        for val, label in [
            ("file_sheet", "File + Sheet / 文件名+表名"),
            ("sheet",      "Sheet Only / 仅表名"),
        ]:
            tk.Radiobutton(self.mode_b_row, text=label,
                           variable=self.sheet_naming, value=val,
                           bg=C["surface"], fg=C["txt_sec"],
                           activebackground=C["surface"],
                           font=F["small"], cursor="hand2",
                           selectcolor=C["surface"],
                           relief="flat").pack(side="left", padx=4)

        # Mode C sub-options
        self.mode_c_row = tk.Frame(inner, bg=C["surface"])
        self.mode_c_row.grid(row=5, column=0, columnspan=4, sticky="w",
                             padx=20, pady=4)
        tk.Label(self.mode_c_row, text="Group By Column / 分组列名:",
                 font=F["small"], bg=C["surface"],
                 fg=C["txt_sec"]).pack(side="left", padx=(0, 10))
        self._styled_entry(self.mode_c_row, self.group_column,
                           width=22).pack(side="left")
        tk.Label(self.mode_c_row,
                 text="  (exact column header / 精确列标题)",
                 font=F["caption"], bg=C["surface"],
                 fg=C["txt_mute"]).pack(side="left")

        # ── divider ──
        HRule(inner).grid(row=6, column=0, columnspan=4, sticky="ew",
                          pady=(8, 12))

        # ── Merge Direction ──
        tk.Label(inner, text="Merge Direction  /  合并方向",
                 font=F["heading"], bg=C["surface"],
                 fg=C["txt"]).grid(row=7, column=0, columnspan=4,
                                   sticky="w", pady=(0, 8))

        dir_data = [
            ("vertical",   "Vertical — Append Rows",
                           "纵向 — 追加行"),
            ("horizontal", "Horizontal — Join Columns",
                           "横向 — 合并列"),
        ]
        for i, (val, en, zh) in enumerate(dir_data):
            row = 8 + i
            rb = tk.Radiobutton(inner,
                                text=f"{en}  /  {zh}",
                                variable=self.direction, value=val,
                                bg=C["surface"], fg=C["txt"],
                                activebackground=C["surface"],
                                font=F["body"], cursor="hand2",
                                selectcolor=C["surface"],
                                relief="flat",
                                command=self._on_direction_change)
            rb.grid(row=row, column=0, columnspan=4, sticky="w", pady=3)

        # Horizontal sub-options
        self.horiz_row = tk.Frame(inner, bg=C["surface"])
        self.horiz_row.grid(row=10, column=0, columnspan=4, sticky="w",
                            padx=20, pady=4)
        tk.Label(self.horiz_row, text="Key Column / 关联列:",
                 font=F["small"], bg=C["surface"],
                 fg=C["txt_sec"]).pack(side="left", padx=(0, 8))
        self._styled_entry(self.horiz_row, self.key_column,
                           width=18).pack(side="left", padx=(0, 16))
        tk.Label(self.horiz_row, text="Join Type / 连接方式:",
                 font=F["small"], bg=C["surface"],
                 fg=C["txt_sec"]).pack(side="left", padx=(0, 8))
        join_cb = ttk.Combobox(self.horiz_row, textvariable=self.join_type,
                               values=["outer  (all rows)",
                                       "left   (left rows)",
                                       "inner  (matching only)"],
                               width=20, state="readonly", font=F["small"])
        join_cb.pack(side="left")
        # map display values to pandas how values
        self._join_map = {
            "outer  (all rows)":       "outer",
            "left   (left rows)":      "left",
            "inner  (matching only)":  "inner",
        }
        self.join_type.set("outer  (all rows)")

    def _refresh_mode_dots(self, inner, mode_data):
        """Update ◉/○ radio dots when mode changes."""
        for row, (val, _en, _zh) in enumerate(mode_data, start=1):
            dot = inner.grid_slaves(row=row, column=0)
            if dot:
                dot[0].config(
                    text="◉" if self.merge_mode.get() == val else "○"
                )

    def _on_mode_change(self, *_):
        mode = self.merge_mode.get()
        if hasattr(self, "_mode_inner"):
            self._refresh_mode_dots(self._mode_inner, self._mode_data)
        if hasattr(self, "mode_b_row"):
            if mode == "B":
                self.mode_b_row.grid()
            else:
                self.mode_b_row.grid_remove()
        if hasattr(self, "mode_c_row"):
            if mode == "C":
                self.mode_c_row.grid()
            else:
                self.mode_c_row.grid_remove()

    def _on_direction_change(self, *_):
        if hasattr(self, "horiz_row"):
            if self.direction.get() == "horizontal":
                self.horiz_row.grid()
            else:
                self.horiz_row.grid_remove()

    # ─── §3 Output Settings ──────────────────────────────────────────────────

    def _build_output_section(self, parent):
        SectionLabel(parent, "③", "Output Settings  /  输出设置").pack(
            fill="x", padx=16, pady=(4, 6))

        card = Card(parent)
        card.pack(fill="x", padx=16, pady=(0, 10))

        grid = tk.Frame(card, bg=C["surface"])
        grid.pack(fill="x", padx=14, pady=12)
        grid.columnconfigure(1, weight=1)

        # Save location
        tk.Label(grid, text="Save To  /  保存位置:",
                 font=F["body"], bg=C["surface"],
                 fg=C["txt"]).grid(row=0, column=0, sticky="w", pady=5, padx=(0,10))

        dir_row = tk.Frame(grid, bg=C["surface"])
        dir_row.grid(row=0, column=1, columnspan=2, sticky="ew")
        dir_row.columnconfigure(0, weight=1)

        self._styled_entry(dir_row, self.output_dir,
                           width=36).grid(row=0, column=0, sticky="ew", padx=(0,6))
        FlatButton(dir_row, "Browse / 浏览",
                   command=self._browse_output_dir).grid(row=0, column=1)

        # File name
        tk.Label(grid, text="File Name  /  文件名:",
                 font=F["body"], bg=C["surface"],
                 fg=C["txt"]).grid(row=1, column=0, sticky="w", pady=5)

        name_row = tk.Frame(grid, bg=C["surface"])
        name_row.grid(row=1, column=1, columnspan=2, sticky="w")
        self._styled_entry(name_row, self.output_name,
                           width=30).pack(side="left")
        tk.Label(name_row, text=".xlsx",
                 font=F["body"], bg=C["surface"],
                 fg=C["txt_mute"]).pack(side="left", padx=4)

        # hint
        tk.Label(grid,
                 text="Existing files will be auto-renamed to avoid overwriting.  "
                      "/  重名文件将自动重命名。",
                 font=F["caption"], bg=C["surface"],
                 fg=C["txt_mute"]).grid(row=2, column=0, columnspan=3,
                                        sticky="w", pady=(0, 4))

    # ─── §4 Actions + Progress ───────────────────────────────────────────────

    def _build_action_section(self, parent):
        act = tk.Frame(parent, bg=C["bg"])
        act.pack(fill="x", padx=16, pady=(4, 12))

        self.start_btn = AccentButton(
            act,
            text="▶   Start Merging  /  开始合并",
            command=self._start_merge
        )
        self.start_btn.pack(side="left")

        FlatButton(act, "↺  Reset / 重置",
                   command=self._reset,
                   padx=14, pady=8).pack(side="left", padx=10)

        # progress card
        prog_card = Card(parent)
        prog_card.pack(fill="x", padx=16, pady=(0, 16))

        self.prog_top = tk.Frame(prog_card, bg=C["surface"])
        self.prog_top.pack(fill="x", padx=12, pady=(10, 4))

        self.prog_status = tk.Label(
            self.prog_top, text="Ready  /  就绪",
            font=F["small"], bg=C["surface"], fg=C["txt_mute"]
        )
        self.prog_status.pack(side="left")

        self.prog_pct = tk.Label(
            self.prog_top, text="",
            font=F["small"], bg=C["surface"], fg=C["accent"]
        )
        self.prog_pct.pack(side="right")

        self.progress_bar = ttk.Progressbar(
            prog_card, style="Horizontal.TProgressbar",
            mode="determinate", length=400
        )
        self.progress_bar.pack(fill="x", padx=12, pady=(0, 10))

    # ─── Log Panel ───────────────────────────────────────────────────────────

    def _build_log_panel(self, parent):
        header = tk.Frame(parent, bg=C["bg"])
        header.pack(fill="x", padx=12, pady=(16, 6))
        tk.Label(header, text="Log  /  日志",
                 font=F["heading"], bg=C["bg"],
                 fg=C["txt"]).pack(side="left")
        FlatButton(header, "Clear / 清空",
                   command=self._clear_log,
                   padx=8, pady=4,
                   font=F["small"]).pack(side="right")

        log_card = Card(parent)
        log_card.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.log_text = tk.Text(
            log_card,
            wrap="word",
            bg=C["log_bg"],
            fg=C["txt"],
            font=F["mono"],
            relief="flat",
            bd=0,
            state="disabled",
            padx=10, pady=8,
            cursor="arrow",
            selectbackground=C["selection"],
        )
        log_vsb = ttk.Scrollbar(log_card, orient="vertical",
                                command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_vsb.set)
        log_vsb.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)

        # coloured log tags
        self.log_text.tag_configure("success", foreground=C["success"])
        self.log_text.tag_configure("error",   foreground=C["error"])
        self.log_text.tag_configure("warning", foreground=C["warning"])
        self.log_text.tag_configure("info",    foreground=C["accent"])
        self.log_text.tag_configure("muted",   foreground=C["txt_mute"])
        self.log_text.tag_configure("dim",     foreground=C["txt_sec"])

    # ─── helpers ─────────────────────────────────────────────────────────────

    def _styled_entry(self, parent, textvariable, width=24) -> tk.Entry:
        """Flat entry with border."""
        e = tk.Entry(
            parent,
            textvariable=textvariable,
            width=width,
            font=F["body"],
            bg=C["surface"],
            fg=C["txt"],
            relief="flat",
            bd=0,
            insertbackground=C["txt"],
            highlightthickness=1,
            highlightbackground=C["border"],
            highlightcolor=C["accent"],
        )
        return e

    # ─── file management ─────────────────────────────────────────────────────

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Excel / CSV Files",
            filetypes=[
                ("Excel & CSV files", "*.xlsx *.xls *.csv"),
                ("Excel files",       "*.xlsx *.xls"),
                ("CSV files",         "*.csv"),
                ("All files",         "*.*"),
            ]
        )
        for p in paths:
            self._register_file(p)

    def _add_folder(self):
        folder = filedialog.askdirectory(title="Select Folder / 选择文件夹")
        if not folder:
            return
        exts = {".xlsx", ".xls", ".csv"}
        pattern = "**/*" if self.include_sub.get() else "*"
        found = 0
        for p in Path(folder).glob(pattern):
            if p.is_file() and p.suffix.lower() in exts:
                self._register_file(str(p))
                found += 1
        if found == 0:
            messagebox.showinfo(
                "No Files Found  /  未找到文件",
                "No Excel/CSV files found in the selected folder.\n"
                "所选文件夹中未找到 Excel/CSV 文件。"
            )

    def _register_file(self, path: str):
        """Add one file to the list (dedup check)."""
        if path in self.file_list:
            return
        try:
            size_b = os.path.getsize(path)
            size_str = (
                f"{size_b/1024:.1f} KB" if size_b < 1_048_576
                else f"{size_b/1_048_576:.1f} MB"
            )
            # peek sheet count without loading data
            ext = Path(path).suffix.lower()
            if ext in (".xlsx", ".xls"):
                try:
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    n_sheets = len(wb.sheetnames)
                    wb.close()
                except Exception:
                    n_sheets = "?"
            else:
                n_sheets = 1

            self.file_list.append(path)
            self.file_tree.insert(
                "", "end",
                values=(Path(path).name, n_sheets, size_str, path)
            )
            self._refresh_file_count()
        except Exception as exc:
            messagebox.showwarning("Cannot Add File / 无法添加文件",
                                   f"{Path(path).name}\n{exc}")

    def _remove_selected(self):
        for iid in self.file_tree.selection():
            path = self.file_tree.item(iid, "values")[3]
            if path in self.file_list:
                self.file_list.remove(path)
            self.file_tree.delete(iid)
        self._refresh_file_count()

    def _clear_files(self):
        self.file_list.clear()
        for iid in self.file_tree.get_children():
            self.file_tree.delete(iid)
        self._refresh_file_count()

    def _refresh_file_count(self):
        n = len(self.file_list)
        if n == 0:
            self.file_count_lbl.config(text="No files added  /  未添加文件")
        else:
            total_b = sum(os.path.getsize(f) for f in self.file_list
                          if os.path.exists(f))
            total_str = (f"{total_b/1024:.0f} KB" if total_b < 1_048_576
                         else f"{total_b/1_048_576:.1f} MB")
            self.file_count_lbl.config(
                text=f"{n} file(s)  /  {n} 个文件  ·  {total_str}"
            )

    # ─── output directory ────────────────────────────────────────────────────

    def _browse_output_dir(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_dir.set(folder)

    # ─── log helpers ─────────────────────────────────────────────────────────

    def _log(self, msg: str, tag: str = "dim"):
        """Thread-safe log append."""
        def _insert():
            self.log_text.configure(state="normal")
            self.log_text.insert("end", msg + "\n", tag)
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.after(0, _insert)

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    # ─── progress ────────────────────────────────────────────────────────────

    def _update_progress(self, current: int, total: int, filename: str = ""):
        pct = int(current / total * 100) if total else 0

        def _do():
            self.progress_bar["value"] = pct
            self.prog_pct.config(text=f"{pct}%")
            self.prog_status.config(
                text=filename if filename else "Processing…  /  处理中…",
                fg=C["txt_sec"]
            )
        self.after(0, _do)

    def _reset_progress(self):
        self.progress_bar["value"] = 0
        self.prog_pct.config(text="")
        self.prog_status.config(text="Ready  /  就绪", fg=C["txt_mute"])

    # ─── validation ──────────────────────────────────────────────────────────

    def _validate(self) -> bool:
        """Returns True if inputs are valid, shows messagebox on error."""
        if not self.file_list:
            messagebox.showwarning(
                "No Files  /  无文件",
                "Please add at least one file before merging.\n"
                "请先添加至少一个文件。"
            )
            return False

        if not self.output_dir.get().strip():
            messagebox.showwarning(
                "No Output Folder  /  无输出目录",
                "Please select an output folder.\n请选择输出文件夹。"
            )
            return False

        if not os.path.isdir(self.output_dir.get().strip()):
            messagebox.showwarning(
                "Invalid Folder  /  文件夹无效",
                f"Output folder does not exist:\n{self.output_dir.get()}\n"
                "输出文件夹不存在，请重新选择。"
            )
            return False

        if not self.output_name.get().strip():
            messagebox.showwarning(
                "No File Name  /  无文件名",
                "Please enter a name for the output file.\n"
                "请输入输出文件名。"
            )
            return False

        if self.merge_mode.get() == "C" and not self.group_column.get().strip():
            messagebox.showwarning(
                "Missing Column  /  缺少列名",
                "Mode C requires a group-by column name.\n"
                "模式 C 需要填写分组列名。"
            )
            return False

        if (self.direction.get() == "horizontal"
                and not self.key_column.get().strip()):
            messagebox.showwarning(
                "Missing Key Column  /  缺少关联列",
                "Horizontal merge requires a key column name.\n"
                "横向合并需要填写关联列名。"
            )
            return False

        # large file warning
        total_mb = sum(os.path.getsize(f) for f in self.file_list
                       if os.path.exists(f)) / 1_048_576
        if total_mb > LARGE_FILE_WARN_MB:
            ok = messagebox.askyesno(
                "Large Files  /  文件较大",
                f"Total input size is {total_mb:.1f} MB.\n"
                "This may take a while. Continue?\n\n"
                f"总大小约 {total_mb:.1f} MB，处理时间可能较长。\n是否继续？"
            )
            if not ok:
                return False

        return True

    # ─── reset ───────────────────────────────────────────────────────────────

    def _reset(self):
        if self._running:
            return
        self._clear_files()
        self._clear_log()
        self._reset_progress()
        self.output_name.set("merged_output")
        self.merge_mode.set("A")
        self.direction.set("vertical")
        self.group_column.set("")
        self.key_column.set("")
        self._on_mode_change()
        self._on_direction_change()
        self._log("App reset.  /  已重置。", "muted")

    # ─── merge execution ─────────────────────────────────────────────────────

    def _start_merge(self):
        if self._running:
            return
        if not self._validate():
            return
        self._running = True
        self.start_btn.set_state(False)
        self._reset_progress()
        self._clear_log()
        threading.Thread(target=self._run_merge, daemon=True).start()

    def _run_merge(self):
        """Runs in a background thread – all UI updates via self.after()."""
        try:
            self._log("━" * 42, "info")
            self._log("Starting merge  /  开始合并", "info")
            self._log("━" * 42, "info")
            self._log(
                f"Files: {len(self.file_list)}  │  "
                f"Mode: {self.merge_mode.get()}  │  "
                f"Direction: {self.direction.get()}",
                "dim"
            )
            self._log("", "dim")

            dp = DataProcessor()
            frames_data = dp.load_frames(
                self.file_list,
                progress_cb=self._update_progress,
                log_cb=self._log
            )

            if not frames_data:
                raise RuntimeError(
                    "No valid data was loaded from the selected files.\n"
                    "所选文件中没有可用数据。"
                )

            mode      = self.merge_mode.get()
            direction = self.direction.get()
            output_sheets: dict[str, pd.DataFrame] = {}

            self._log("", "dim")
            # ── Mode A ──────────────────────────────────────────────────────
            if mode == "A":
                self._log("Mode A: Merging all data into one sheet…", "info")
                all_dfs = [df for _, _, df in frames_data]

                if direction == "vertical":
                    merged = DataProcessor.merge_vertical(all_dfs)
                    self._log(
                        f"  ✓  Vertically merged: "
                        f"{len(merged):,} rows × {len(merged.columns)} cols",
                        "success"
                    )
                else:
                    raw_how = self.join_type.get()
                    how = self._join_map.get(raw_how, "outer")
                    key = self.key_column.get().strip()
                    # validate key exists in every sheet
                    for fp, sn, df in frames_data:
                        if key not in df.columns:
                            raise ValueError(
                                f"Key column '{key}' not found in\n"
                                f"'{Path(fp).name}' → sheet '{sn}'.\n\n"
                                f"关联列 '{key}' 在文件 '{Path(fp).name}' "
                                f"的工作表 '{sn}' 中不存在。"
                            )
                    merged = DataProcessor.merge_horizontal(all_dfs, key, how)
                    self._log(
                        f"  ✓  Horizontally merged ({how}): "
                        f"{len(merged):,} rows × {len(merged.columns)} cols",
                        "success"
                    )
                output_sheets["Merged"] = merged

            # ── Mode B ──────────────────────────────────────────────────────
            elif mode == "B":
                self._log("Mode B: Creating separate sheets…", "info")
                naming = self.sheet_naming.get()
                used: list[str] = []
                for fp, sname, df in frames_data:
                    if naming == "file_sheet":
                        stem = Path(fp).stem[:15]
                        raw  = f"{stem}_{sname}"
                    else:
                        raw = sname
                    safe = DataProcessor.safe_sheet_name(raw, used)
                    used.append(safe)
                    output_sheets[safe] = df
                    self._log(
                        f"  ✓  Sheet '{safe}': {len(df):,} rows",
                        "success"
                    )
                self._log(
                    f"\n  Total sheets: {len(output_sheets)}", "info"
                )

            # ── Mode C ──────────────────────────────────────────────────────
            elif mode == "C":
                col = self.group_column.get().strip()
                self._log(f"Mode C: Grouping by column '{col}'…", "info")

                # check column exists everywhere
                for fp, sn, df in frames_data:
                    if col not in df.columns:
                        raise ValueError(
                            f"Column '{col}' not found in\n"
                            f"'{Path(fp).name}' → sheet '{sn}'.\n\n"
                            f"列 '{col}' 在文件 '{Path(fp).name}' "
                            f"的工作表 '{sn}' 中不存在。"
                        )

                combined = DataProcessor.merge_vertical(
                    [df for _, _, df in frames_data]
                )
                used: list[str] = []
                groups = list(combined.groupby(col, sort=True))
                for i, (gval, gdf) in enumerate(groups):
                    safe = DataProcessor.safe_sheet_name(str(gval)[:28], used)
                    used.append(safe)
                    output_sheets[safe] = gdf.reset_index(drop=True)
                    self._log(
                        f"  ✓  '{safe}': {len(gdf):,} rows",
                        "success"
                    )
                    self._update_progress(i + 1, len(groups), str(gval))
                self._log(
                    f"\n  Total groups: {len(output_sheets)}", "info"
                )

            # ── Save ────────────────────────────────────────────────────────
            out_path = self._resolve_output_path()
            self._log(f"\nSaving → {out_path}", "info")
            self._save_xlsx(output_sheets, out_path)

            self._log("", "dim")
            self._log("━" * 42, "success")
            self._log(
                f"✓  Done!  /  完成！  ({len(output_sheets)} sheet(s))",
                "success"
            )
            self._log(f"   {out_path}", "success")
            self._log("━" * 42, "success")

            self._update_progress(1, 1, "Complete!  /  完成！")
            self.after(0, lambda: messagebox.showinfo(
                "Success  /  成功",
                f"Merge complete!  Saved to:\n{out_path}\n\n"
                f"合并完成！文件已保存至:\n{out_path}"
            ))

        except Exception as exc:
            self._log("", "dim")
            self._log("━" * 42, "error")
            self._log(f"✗  Error  /  错误: {exc}", "error")
            self._log("━" * 42, "error")
            self._log("", "dim")
            self._log(traceback.format_exc(), "muted")
            self.after(0, lambda e=exc: messagebox.showerror(
                "Error  /  错误",
                f"Merge failed:\n{e}\n\n合并失败，请查看日志了解详情。"
            ))

        finally:
            self._running = False
            self.after(0, lambda: self.start_btn.set_state(True))

    # ─── output helpers ──────────────────────────────────────────────────────

    def _resolve_output_path(self) -> str:
        """Build output path; auto-rename if file already exists."""
        folder = self.output_dir.get().strip()
        name   = self.output_name.get().strip()
        if not name.lower().endswith(".xlsx"):
            name += ".xlsx"
        path = os.path.join(folder, name)
        if not os.path.exists(path):
            return path
        base = name[:-5]  # strip .xlsx
        for i in range(1, 10_000):
            candidate = os.path.join(folder, f"{base}_{i}.xlsx")
            if not os.path.exists(candidate):
                self._log(
                    f"  ⚠  File exists – renamed to {Path(candidate).name}",
                    "warning"
                )
                return candidate
        return path  # last resort

    def _save_xlsx(self, sheets: dict[str, pd.DataFrame], path: str):
        """Write all sheets to an .xlsx file with auto column widths."""
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                # ── auto column widths ──
                for col_cells in ws.columns:
                    max_len = 0
                    for cell in col_cells:
                        try:
                            cl = len(str(cell.value)) if cell.value is not None else 0
                            if cl > max_len:
                                max_len = cl
                        except Exception:
                            pass
                    adj = min(max_len + 4, 60)
                    ws.column_dimensions[
                        col_cells[0].column_letter
                    ].width = max(adj, 8)

                # ── freeze header row ──
                ws.freeze_panes = "A2"

                # ── style header row ──
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(
                    start_color="EBF3FD", end_color="EBF3FD",
                    fill_type="solid"
                )
                header_font = Font(bold=True, color="1A1A1A")
                for cell in ws[1]:
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )

        self._log(f"  ✓  Saved ({os.path.getsize(path)/1024:.1f} KB)", "success")


# ═════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

def main():
    app = ExcelMergerApp()
    # nice window icon (uses tk's built-in photo support)
    try:
        img = tk.PhotoImage(data=(
            "R0lGODlhEAAQAIAAAAAAAP///yH5BAAAAAAALAAAAAAQABAAAAIjhI+py+0Po5y02ouz"
            "3rz7D4biSJbmiabqyrbuC8fyTNf2UQAAOw=="
        ))
        app.iconphoto(True, img)
    except Exception:
        pass
    app.mainloop()


if __name__ == "__main__":
    main()
